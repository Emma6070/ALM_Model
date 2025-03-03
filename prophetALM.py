import pandas as pd
from prophet import Prophet
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from datetime import datetime, timedelta

def create_sheet(wb, sheet_name, headers, data):
    """Create and fill a worksheet"""
    sheet = wb.create_sheet(title=sheet_name)
    sheet.append(headers)
    for row in dataframe_to_rows(data, index=False, header=False):
        sheet.append(row)
    return sheet

def calculate_bel(cash_flows, discount_rate, periods):
    """Calculate Best Estimate Liability"""
    return sum(cf / (1 + discount_rate)**t for t, cf in enumerate(cash_flows[:periods], 1))

def calculate_risk_margin(impact, probability, risk_weight):
    """Calculate risk margin based on risk factors"""
    return impact * probability * risk_weight

# Load historical asset values (replace with your actual data)
historical_data = {
    'ds': pd.date_range(start='2023-01-01', periods=12, freq='M'),
    'y': [1000000, 1020000, 1040000, 1065000, 1080000, 1100000,
          1120000, 1140000, 1160000, 1180000, 1200000, 1220000]
}
df = pd.DataFrame(historical_data)

# Initialize and fit Prophet model
model = Prophet(yearly_seasonality=True)
model.fit(df)

# Make future predictions (12 months ahead)
future = model.make_future_dataframe(periods=12, freq='M')
forecast = model.predict(future)

# Visualize the forecast
fig = model.plot(forecast)
plt.title('Asset Value Forecast')
plt.show()

# Create new workbook
workbook = Workbook()

# 1. Inputs sheet - ALM assumptions
inputs_headers = [
    'Parameter', 'Value'
]
inputs_data = pd.DataFrame({
    'Parameter': ['Initial Assets', 'Discount Rate', 'Risk-Free Rate', 
                 'Credit Spread', 'Liability Duration', 'Target Funding Ratio'],
    'Value': [1000000, 0.03, 0.02, 0.01, 10, 1.10]
})
create_sheet(workbook, 'Assumptions', inputs_headers, inputs_data)

# 2. Assets sheet
forecast['ds'] = pd.to_datetime(forecast['ds']).dt.strftime('%Y-%m-%d')
assets_headers = ['Date', 'Market Value', 'Expected Return', 'Duration', 'Credit Quality']
assets_data = forecast[['ds', 'yhat']].copy()
assets_data.columns = ['Date', 'Market Value']
assets_data['Expected Return'] = 0.05
assets_data['Duration'] = 8
assets_data['Credit Quality'] = 'AA'
create_sheet(workbook, 'Assets', assets_headers, assets_data)

# 3. Liabilities sheet
liability_headers = ['Date', 'Expected Cashflow', 'Duration', 'Best Estimate']
liability_data = forecast[['ds']].copy()
liability_data['Expected Cashflow'] = 50000  # Annual expected claims
liability_data['Duration'] = 10
liability_data['Best Estimate'] = calculate_bel(
    [50000]*12, discount_rate=0.03, periods=10
)
create_sheet(workbook, 'Liabilities', liability_headers, liability_data)

# 4. Risk Assessment sheet
risk_headers = ['Risk Type', 'Impact', 'Probability', 'Risk Weight', 'Risk Margin']
risk_types = ['Market Risk', 'Credit Risk', 'Insurance Risk', 'Operational Risk']
risk_data = pd.DataFrame({
    'Risk Type': risk_types,
    'Impact': [5000000, 3000000, 2000000, 1000000],
    'Probability': [0.05, 0.03, 0.04, 0.02],
    'Risk Weight': [1.5, 1.3, 1.2, 1.1],
    'Risk Margin': 0.0
})
risk_data['Risk Margin'] = risk_data.apply(
    lambda x: calculate_risk_margin(x['Impact'], x['Probability'], x['Risk Weight']), 
    axis=1
)
create_sheet(workbook, 'Risk Assessment', risk_headers, risk_data)

# 5. ALM Metrics sheet
metrics_headers = ['Metric', 'Value']
funding_ratio = assets_data['Market Value'].mean() / liability_data['Best Estimate'].mean()
duration_gap = assets_data['Duration'].mean() - liability_data['Duration'].mean()

metrics_data = pd.DataFrame({
    'Metric': ['Funding Ratio', 'Duration Gap', 'Total Risk Margin'],
    'Value': [funding_ratio, duration_gap, risk_data['Risk Margin'].sum()]
})
create_sheet(workbook, 'ALM Metrics', metrics_headers, metrics_data)

# Save the workbook
excel_file = 'ALM_model.xlsx'
workbook.save(excel_file)
print(f"ALM model saved to {excel_file}")
